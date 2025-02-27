import os
import openpyxl
import requests
import configparser
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime

from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties



now = datetime.now().strftime("%Y-%m-%d_%Hh%M")  # ✅ Définition 

# 📌 Charger la configuration depuis pybliometrics.cfg
config = configparser.ConfigParser()
config.read(r"C:\Users\alima\.config\pybliometrics.cfg") # Sur Windows
# config.read(r"/Users/alexandreleao/.config/pybliometrics.cfg") # Sur MacOs

API_KEY = config["Authentication"]["APIKey"]
INST_TOKEN = config["Authentication"]["InstToken"]

# ✅ Vérifier si l'API Key et le Token sont bien récupérés
if not API_KEY or not INST_TOKEN:
    print("\u274c Erreur : Impossible de récupérer l'API Key ou le Token Institutionnel depuis pybliometrics.cfg")
    exit()

print(f"\u2705 API Key récupérée : {API_KEY[:5]}**")
print(f"\u2705 Institution Token récupéré : {INST_TOKEN[:5]}**\n")

# ✅ Demander l'ID de l’auteur
AUTHOR_ID = input("Entrez l'ID Scopus de l’auteur : ").strip()

# ✅ Obtenir l'année actuelle
current_year = datetime.now().year

# ✅ Demander à l'utilisateur de saisir la plage d'années avec une option par défaut
start_year_input = input(f"Entrez l'année de début (ou appuyez sur Entrée pour {current_year - 6}) : ").strip()
end_year_input = input(f"Entrez l'année de fin (ou appuyez sur Entrée pour {current_year}) : ").strip()

# ✅ Vérifier et attribuer les valeurs par défaut si l'utilisateur appuie sur Entrée
if start_year_input and not start_year_input.isdigit():
    print("\u274c Erreur : L'année de début doit être un nombre entier.")
    exit()
if end_year_input and not end_year_input.isdigit():
    print("\u274c Erreur : L'année de fin doit être un nombre entier.")
    exit()

start_year = int(start_year_input) if start_year_input else current_year - 6
end_year = int(end_year_input) if end_year_input else current_year

# ✅ Vérifier que l'année de fin est postérieure ou égale à l'année de début
if start_year > end_year:
    print("\u274c Erreur : L'année de début doit être inférieure ou égale à l'année de fin.")
    exit()

print(f"📊 Plage d'années sélectionnée : {start_year} - {end_year}")



if not AUTHOR_ID:
    print("\u274c Erreur : L'ID de l'auteur ne peut pas être vide.")
    exit()


def display_subject_areas(df):
    print("\nCode ASJC\tSubject Area\tScholarly Output")
    for index, row in df.iterrows():
        # Utilisez les noms de colonnes corrects ici
        print(f"{row['ASJC Code']}\t{row['Subject Area']}\t{row['Scholarly Output']}")



# ✅ Définition des headers AVANT d'effectuer le premier appel API
headers = {
    "X-ELS-APIKey": API_KEY,
    "X-ELS-Insttoken": INST_TOKEN,
    "Accept": "application/json"
}

# ✅ Définition de l'URL de l'API Scopus pour récupérer le total des publications sur la plage choisie
SCOPUS_TOTAL_URL = "https://api.elsevier.com/content/search/scopus"
params_total = {
    "query": f"AU-ID({AUTHOR_ID}) AND PUBYEAR > {start_year-1} AND PUBYEAR < {end_year+1}",
    "count": "0"
}


response_total = requests.get(SCOPUS_TOTAL_URL, headers=headers, params=params_total)

if response_total.status_code == 200:
    total_data = response_total.json()
    try:
        total_publications_plage = int(total_data["search-results"]["opensearch:totalResults"])
        print(f"\u2705 Nombre total de publications ({start_year}-{end_year}) : {total_publications_plage}")
    except KeyError:
        print("\u26a0\ufe0f Impossible de récupérer le nombre total de publications sur la période demandée.")
        total_publications_plage = None
else:
    print(f"\u274c Erreur API Scopus : {response_total.status_code} - {response_total.text}")
    total_publications_plage = None



response_total = requests.get(SCOPUS_TOTAL_URL, headers=headers, params=params_total)

if response_total.status_code == 200:
    total_data = response_total.json()
    try:
        total_publications = int(total_data["search-results"]["opensearch:totalResults"])
        print(f"\u2705 Nombre total de publications (2019-2025) : {total_publications}")
    except KeyError:
        print("\u26a0\ufe0f Impossible de récupérer le nombre total de publications. Utilisation de la somme des Subject Areas.")
        total_publications = None
else:
    print(f"\u274c Erreur API Scopus : {response_total.status_code} - {response_total.text}")
    total_publications = None

# ✅ Définition de l'URL de l'API Scopus pour les Subject Areas
params = {
    "query": f"AU-ID({AUTHOR_ID}) AND PUBYEAR > 2018 AND PUBYEAR < 2026",
    "facets": "subjarea",
    "count": "0"
}

response = requests.get(SCOPUS_TOTAL_URL, headers=headers, params=params)
print(response.json())  # Ceci affichera la structure JSON de la réponse

if response.status_code != 200:
    print(f"\u274c Erreur API Scopus : {response.status_code} - {response.text}")
    exit()

data = response.json()
if response.status_code == 200:
    data = response.json()

    # Extraire les codes ASJC et les noms des domaines
    entries = data["search-results"]["entry"]
    subjareas = []
    for entry in entries:
        if 'subjareas' in entry:  # S'assurer que 'subjareas' est présent
            subjareas.extend(entry['subjareas'])

    # Créer un DataFrame à partir des subjareas extraits
    df = pd.DataFrame(subjareas)
    print(df)  # Vérifier la structure du DataFrame

    # Continuer avec d'autres traitements ou manipulations...
else:
    print(f"Erreur API Scopus : {response.status_code} - {response.text}")
    exit()

if "search-results" in data and "facet" in data["search-results"]:
    subject_areas = data["search-results"]["facet"]["category"]
    df = pd.DataFrame(subject_areas)
    df = df.rename(columns={"label": "Subject Area", "hitCount": "Scholarly Output"})

     # Après avoir chargé les données dans le DataFrame
    if 'some_other_name_for_asjc' in df.columns:
            df.rename(columns={'some_other_name_for_asjc': 'ASJC Code'}, inplace=True)

    df["Scholarly Output"] = df["Scholarly Output"].astype(int)

    
    # Trier par ordre décroissant
    df = df.sort_values(by="Scholarly Output", ascending=False)

    # ✅ Removing "(all)" from subject names
    df["Subject Area"] = df["Subject Area"].str.replace(r"\s*\(all\)", "", regex=True)

    # Calcul des pourcentages corrects
    if total_publications:
        df["Percentage"] = (df["Scholarly Output"] / total_publications) * 100
    else:
        df["Percentage"] = (df["Scholarly Output"] / df["Scholarly Output"].sum()) * 100

    # 📌 Récupérer le nom et prénom de l'auteur AVANT de générer le nom du fichier
    SCOPUS_AUTHOR_URL = f"https://api.elsevier.com/content/author/author_id/{AUTHOR_ID}"

    response = requests.get(SCOPUS_AUTHOR_URL, headers=headers)

    if response.status_code == 200:
        author_data = response.json()  # 📌 Définition correcte de author_data
        try:
            total_publications = int(author_data["author-retrieval-response"][0]["coredata"]["document-count"])
            print(f"✅ Nombre total de publications (2019-2025) : {total_publications}")
        except KeyError:
            print("⚠️ Impossible de récupérer le nombre total de publications. Utilisation de la somme des Subject Areas.")
            total_publications = None
    else:
        print(f"❌ Erreur API Scopus : {response.status_code} - {response.text}")
        total_publications = None


    # ✅ Récupérer le nom et prénom de l'auteur AVANT de générer le nom du fichier
    try:
        author_name = author_data["author-retrieval-response"][0]["author-profile"]["preferred-name"]
        author_firstname = author_name["given-name"]
        author_lastname = author_name["surname"]
    except KeyError:
        print("⚠️ Impossible de récupérer le nom complet de l’auteur. Utilisation de l’ID.")
        author_firstname = "Auteur"
        author_lastname = AUTHOR_ID

    # ✅ Générer le nom du fichier après avoir défini les variables
    filename = f"{now}{author_lastname}{author_firstname.replace(' ', '_')}.xlsx"

    # ✅ Exporter vers Excel
    now = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    filename = f"{now}{author_lastname}{author_firstname.replace(' ', '_')}.xlsx"

    file_path = os.path.join(r"C:\Users\alima\Desktop\AlexStage\Code_graphique_test_project\output", filename) # Sur Windows
    # file_path = os.path.join(r"/Users/alexandreleao/Documents/projectBiblio/output", filename) # Sur MacOs
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Areas"

    # ✅ Ajouter les informations générales en haut du tableau
    ws.append(["Data set", "Publications by Subject Area"])
    ws.append(["Entity", AUTHOR_ID])
    ws.append(["Subject classification", "ASJC"])
    ws.append(["Filtered by", "not filtered"])
    ws.append(["Types of publications included", "all publication types"])
    ws.append(["Self-citations", "included"])
    ws.append(["Data source", "Scopus"])
    ws.append(["Date last updated", "29 January 2025"])  # 📌 À automatiser si nécessaire
    ws.append(["Date exported", datetime.now().strftime("%d %B %Y")])
    ws.append(["Year range", f"{start_year} to {end_year}"])  # 🔹 Afficher la plage d'années choisie
    ws.append(["Total général", total_publications])  # 🔹 Ajouter le total général
    ws.append(["Total dans la période choisie", total_publications_plage])  # 🔹 Afficher le total dynamique
    ws.append([])  # ✅ Ligne vide pour séparer les infos du tableau

    # ✅ Ensure only the required columns are kept
    df = df[["Subject Area", "Scholarly Output", "Percentage"]]

    # ✅ Convert Percentage column to numeric and clean up formatting
    df["Percentage"] = pd.to_numeric(df["Percentage"], errors="coerce")

    # ✅ Add a new column for displaying percentages with "%" symbol
    df["Percentage Display"] = df["Percentage"].apply(lambda x: f"{x:.1f}%")

    # ✅ Format numbers to 1 decimal place
    df["Scholarly Output"] = df["Scholarly Output"].round(1)
    df["Percentage"] = df["Percentage"].round(1)

    # ✅ Sort DataFrame by Scholarly Output (highest to lowest)
    df = df.sort_values(by="Scholarly Output", ascending=False)

    

    # ✅ Import nécessaire AVANT d'utiliser DataLabelList
    from openpyxl.chart.label import DataLabelList  

# Sort the data from highest to lowest
ws.append(["Subject Area", "Scholarly Output", "Percentage Display"])
ws.cell(row=14, column=4, value="Percentage Display %")  # Add header for Column D
sorted_data = sorted(df.values.tolist(), key=lambda x: x[1], reverse=False)

# Insert the sorted data into the Excel sheet
for row in sorted_data:
    ws.append(row)    

# ✅ Dynamically detect the starting and ending row
start_row = ws.max_row - len(df) + 1  # Detects where data starts
end_row = ws.max_row  # Detects last row with data

# Define the data range for the chart (Percentage column)
data_range = Reference(ws, min_col=3, min_row=start_row, max_row=end_row)  # Column C (3)

# ✅ Define the data range for the chart labels (use Column D for Percentage Display)
label_range = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)  # Column D (4)

# Define the categories range for the chart (Subject Area names)
categories_range = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)

from openpyxl.chart.shapes import GraphicalProperties

# Create a horizontal bar chart
chart = BarChart()
chart.type = "bar"  # Horizontal bar chart
chart.title = None
chart.y_axis.title = None
chart.x_axis.title = None

# ✅ Properly remove the external grey frame (border)
if chart.graphical_properties is None:
    chart.graphical_properties = GraphicalProperties()

chart.graphical_properties.line.noFill = True  # ✅ Ensures no border is applied

# Add data and categories to the chart
chart.add_data(data_range, titles_from_data=False)  # Ensure we don't take the column title
chart.set_categories(categories_range)

if chart.series:
    for series in chart.series:
        if series is not None:  # ✅ Ensure series exists before modifying
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True  # ✅ Show percentage values correctly (now from Column D)
            series.dLbls.showCatName = True  # ❌ Hide category names from labels
            series.dLbls.showSerName = False  # ✅ Disable "Series1"
            series.dLbls.position = "outEnd"  # ✅ Position labels at the end of bars
            series.dLbls.numFmt = "0.0%"  # ✅ Ensure 1 decimal place + "%"
            series.dLbls.separator = " "  # ✅ Ensure good spacing between numbers and bars
            series.dLbls.showLeaderLines = True  # ✅ Adds leader lines to separate labels from bars

            # ✅ Correctly set label color to deep blue
            series.dLbls.txPr = RichText(
                p=[ParagraphProperties(defRPr=CharacterProperties(solidFill="1F4E79"))]  # ✅ Dark blue labels
            )


    # ✅ Reduce bar width to make it thinner
    series.gapWidth = 400  # Increase gap width (default is ~150)

    # ✅ Simulate rounded markers by adding an extra small bar
    # ✅ Remove ALL border lines (internal & external)
    series.graphicalProperties.line.noFill = True  # Removes outline
    series.graphicalProperties.solidFill = "1F4E79"  # Keep solid blue fill
    series.spPr.ln.noFill = True  # ✅ Ensures NO external border is applied

    
    
    # ✅ Remove the "Series1" legend
    chart.legend = None  # Completely remove the legend

    # # ✅ Restore bar colors
    # for series in chart.series:
    #     series.graphicalProperties.solidFill = "4472C4"  # Restore blue color

    # ✅ Reverse the Y-axis order to invert labels
    chart.y_axis.reverseOrder = True  # Invert label order on the graph
    chart.y_axis.tickLblPos = "low"  # Align labels properly

    # Style the chart
    chart.style = 10  # Apply a predefined style
    chart.y_axis.majorGridlines = None  # Remove gridlines
    chart.x_axis.majorGridlines = None  # Remove gridlines

    chart.y_axis.majorTickMark = "none"  # ✅ Removes extra tick marks
    chart.y_axis.labelOffset = 0  # ✅ Moves the category labels closer to the bars

    chart.layout = Layout(
    manualLayout=ManualLayout(
        x=0.01,  # Move bars even closer to left
        y=0.05,
        w=0.7,  # Adjust width for better spacing
        h=0.85
    )
)


    # Add the chart to the worksheet
    ws.add_chart(chart, "E10")  # ✅ Déplace le graphique vers le bas pour plus d’espace
    # chart.width = 20  # ✅ Augmente la largeur du graphique
    # chart.height = 10  # ✅ Augmente la hauteur du graphique

    
print(df.columns)

# Après avoir préparé df avec toutes les données nécessaires
display_subject_areas(df)

# Demande à l'utilisateur de choisir les catégories/sous-catégories
selected_codes = input("Entrez les codes ASJC des catégories/sous-catégories à inclure, séparés par une virgule : ")
selected_codes_list = [code.strip() for code in selected_codes.split(',')]
filtered_df = df[df['ASJC Code'].isin(selected_codes_list)]

# Ajout des données filtrées dans la feuille Excel
for row in filtered_df.itertuples():
    ws.append([row.SubjectArea, row.ScholarlyOutput, row.PercentageDisplay])

# Configuration et ajout du graphique
chart = BarChart()
chart.add_data(data=Reference(ws, min_col=2, min_row=2, max_row=ws.max_row, max_col=3), titles_from_data=True)
chart.set_categories(categories=Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
chart.legend = None
chart.y_axis.reverseOrder = True
chart.style = 10
chart.y_axis.majorGridlines = None
chart.x_axis.majorGridlines = None
chart.y_axis.majorTickMark = "none"
chart.y_axis.labelOffset = 0

chart.layout = Layout(
    manualLayout=ManualLayout(
        x=0.01,  # Move bars even closer to left
        y=0.05,
        w=0.7,  # Adjust width for better spacing
        h=0.85
    )
)

ws.add_chart(chart, "E10")

# Enregistrement du fichier
wb.save(file_path)
if not filtered_df.empty:
    print(f"\n\u2705 Données exportées avec succès dans '{filename}'")
else:
    print("\u26a0\ufe0f Aucun regroupement par Subject Area trouvé. Vérifiez l'ID de l'auteur ou l'API Key.")


