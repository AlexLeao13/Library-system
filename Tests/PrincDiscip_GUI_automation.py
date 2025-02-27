import os
import requests
import configparser
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl.drawing.line import LineProperties

# Fichier pour stocker le chemin de sauvegarde des rapports
SAVE_PATH_CONFIG = os.path.expanduser("~/.config/princdiscip_config.ini")

# Fichier pour les clÃ©s API
API_CONFIG_PATH = os.path.expanduser("~/.config/pybliometrics.cfg")



def find_config_file():
    """Recherche le fichier pybliometrics.cfg dans diffÃ©rents emplacements."""
    possible_paths = [
        os.path.expanduser("~/.config/pybliometrics.cfg"),  # Emplacement par dÃ©faut (Linux/Mac/Windows)
        os.path.join(os.getenv("APPDATA", ""), "pybliometrics.cfg"),  # Windows : C:\Users\Nom\AppData\Roaming\
        os.path.join(os.path.expanduser("~"), ".pybliometrics.cfg"),  # Dans le dossier utilisateur
        os.path.join(os.getcwd(), "pybliometrics.cfg")  # Dossier actuel du script
    ]

    for path in possible_paths:
        if os.path.exists(path):
            return path

    return None

import os
import configparser

CONFIG_PATH = os.path.expanduser("~/.config/pybliometrics.cfg")

def load_api_keys():
    """Charger les clÃ©s API Ã  partir du fichier pybliometrics.cfg sans le modifier."""
    if not os.path.exists(API_CONFIG_PATH):
        raise FileNotFoundError(f"Le fichier de configuration {API_CONFIG_PATH} est introuvable.")

    config = configparser.ConfigParser()
    config.read(API_CONFIG_PATH)

    if 'Authentication' not in config:
        raise ValueError("La section [Authentication] est absente du fichier de configuration.")

    try:
        api_key = config['Authentication']['APIKey']
        inst_token = config['Authentication']['InstToken']
        return api_key, inst_token
    except KeyError as e:
        raise ValueError(f"ClÃ© manquante dans la section [Authentication] : {e}")




def load_save_path():
    """Charger le chemin de sauvegarde depuis la configuration locale."""
    config = configparser.ConfigParser()

    if os.path.exists(SAVE_PATH_CONFIG):
        config.read(SAVE_PATH_CONFIG)
        return config.get("Settings", "save_path", fallback=None)
    else:
        return None

def save_save_path(path):
    """Enregistrer le chemin de sauvegarde dans un fichier sÃ©parÃ©."""
    config = configparser.ConfigParser()
    config["Settings"] = {"save_path": path}

    os.makedirs(os.path.dirname(SAVE_PATH_CONFIG), exist_ok=True)

    with open(SAVE_PATH_CONFIG, "w") as configfile:
        config.write(configfile)



#  RÃ©cupÃ©rer les publications par plage d'annÃ©es
def get_publication_count(author_id, start_year, end_year, api_key, inst_token):
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "X-ELS-Insttoken": inst_token, "Accept": "application/json"}
    params = {"query": f"AU-ID({author_id}) AND PUBYEAR > {start_year-1} AND PUBYEAR < {end_year+1}", "count": "0"}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        raise RuntimeError(f"Erreur API : {response.status_code} - {response.text}")
    return int(response.json()["search-results"]["opensearch:totalResults"])

#  RÃ©cupÃ©rer les principales disciplines
def get_subject_areas(author_id, start_year, end_year, api_key, inst_token):
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "X-ELS-Insttoken": inst_token, "Accept": "application/json"}
    params = {"query": f"AU-ID({author_id}) AND PUBYEAR > {start_year-1} AND PUBYEAR < {end_year+1}", "facets": "subjarea", "count": "0"}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        raise RuntimeError(f"Erreur API : {response.status_code} - {response.text}")

    data = response.json()
    subject_areas = data.get("search-results", {}).get("facet", {}).get("category", [])
    if not subject_areas:
        raise ValueError("Aucune catÃ©gorie trouvÃ©e.")

    df = pd.DataFrame(subject_areas)
    df.rename(columns={"label": "Subject Area", "hitCount": "Scholarly Output"}, inplace=True)
    df["Scholarly Output"] = df["Scholarly Output"].astype(int)
    return df.sort_values(by="Scholarly Output", ascending=False)

def get_author_id_by_name(author_name, api_key, inst_token):
    if "," not in author_name:
        raise ValueError("Le nom doit Ãªtre au format : Nom, PrÃ©nom")

    last_name, first_name = [part.strip().title() for part in author_name.split(",", 1)]
    formatted_name = f"{last_name}, {first_name}"

    url = "https://api.elsevier.com/content/search/author"
    headers = {
        "X-ELS-APIKey": api_key,
        "X-ELS-Insttoken": inst_token,
        "Accept": "application/json"
    }
    # Supprimer la virgule pour le format Nom PrÃ©nom
    formatted_name = formatted_name.replace(",", "")
    # SÃ©paration du nom et du prÃ©nom
    last_name, first_name = formatted_name.split(" ", 1)
    params = {"query": f"AUTHLAST({last_name}) AND AUTHFIRST({first_name})", "count": 5}




    #  LOG : Afficher la requÃªte envoyÃ©e
    print(f"\nðŸ” RequÃªte envoyÃ©e : {url}")
    print(f"ParamÃ¨tres : {params}")
    print(f"En-tÃªtes : {headers}\n")

    # Envoyer la requÃªte
    response = requests.get(url, headers=headers, params=params)

    # LOG de la rÃ©ponse brute
    print(f"RÃ©ponse brute de l'API : {response.status_code}")
    print(response.json())

    if response.status_code != 200:
        raise RuntimeError(f"Erreur API : {response.status_code} - {response.json()}")

    results = response.json().get("search-results", {}).get("entry", [])
    if not results:
        raise ValueError(f"Aucun auteur trouvÃ© pour : {formatted_name}")

    # Afficher les auteurs trouvÃ©s
    for entry in results:
        print(f"Nom : {entry.get('preferred-name', {}).get('surname')}, {entry.get('preferred-name', {}).get('given-name')}")
        print(f"ID Scopus : {entry['dc:identifier'].split(':')[-1]}")
        print("---")

    return results[0]["dc:identifier"].split(":")[-1]





#  CrÃ©er un fichier Excel avec le graphique
def create_excel_report(author_id, start_year, end_year, df, total_publications, save_path):
    now = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    filename = f"{now}_Scopus_{author_id}.xlsx"

    # Utiliser le chemin dÃ©fini par l'utilisateur pour sauvegarder le fichier
    output_dir = save_path
    os.makedirs(output_dir, exist_ok=True)
    
    # Chemin complet du fichier
    file_path = os.path.join(output_dir, filename)

    # Calcul des pourcentages corrects
    if total_publications:
        df["Percentage"] = df["Scholarly Output"] / total_publications
    else:
        df["Percentage"] = df["Scholarly Output"] / df["Scholarly Output"].sum()


    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Areas"

    #  Ajouter les informations gÃ©nÃ©rales en haut du tableau
    ws.append(["Data set", "Publications by Subject Area"])
    ws.append(["Entity", author_id])
    ws.append(["Subject classification", "ASJC"])
    ws.append(["Filtered by", "not filtered"])
    ws.append(["Types of publications included", "all publication types"])
    ws.append(["Self-citations", "included"])
    ws.append(["Data source", "Scopus"])
    ws.append(["Date last updated", "29 January 2025"])  #  Ã€ automatiser si nÃ©cessaire
    ws.append(["Date exported", datetime.now().strftime("%d %B %Y")])
    ws.append(["Year range", f"{start_year} to {end_year}"])  # ðŸ”¹ Afficher la plage d'annÃ©es choisie
    ws.append(["Total gÃ©nÃ©ral", total_publications])  # ðŸ”¹ Ajouter le total gÃ©nÃ©ral
    ws.append(["Total dans la pÃ©riode choisie", total_publications])  # ðŸ”¹ Afficher le total dynamique
    ws.append([])  #  Ligne vide pour sÃ©parer les infos du tableau


    # Retirer "(all)" des noms des Subject Areas
    df["Subject Area"] = df["Subject Area"].str.replace(r"\s*\(all\)", "", regex=True)

   
    # Ajouter les donnÃ©es principales
    ws.append(["Subject Area", "Scholarly Output", "Percentage", "GraphPercentage"])  # Ajouter les en-tÃªtes
   
   
     #  Sort DataFrame by Scholarly Output (highest to lowest)
    df = df.sort_values(by="Scholarly Output", ascending=False)
    df = df.iloc[::-1]  # Inverser l'ordre des lignes

    # CrÃ©er une colonne pour le graphique avec les pourcentages multipliÃ©s par 100 (sans affecter le tableau)
    df["GraphPercentage"] = df["Percentage"]   # Ajustement pour l'affichage correct du graphique



    for _, row in df.iterrows():
        #ws.append([row["Subject Area"], row["Scholarly Output"], row["Percentage"], round(row["GraphPercentage"], 1)])
        ws.append([row["Subject Area"], row["Scholarly Output"], row["Percentage"], row["GraphPercentage"]])
    
   
    #  Dynamically detect the starting and ending row
    start_row = ws.max_row - len(df) + 1  # Detects where data starts
    end_row = ws.max_row  # Detects last row with data

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '0.0%'

    # CrÃ©er le graphique
    chart = BarChart()
    chart.type = "bar"  # This creates a horizontal bar chart
    chart.style = 10
    # chart.y_axis.reverseOrder = True  #  Reverse the order of bars (longest bar at the top)
    chart.y_axis.labelOffset = 0

    chart.title = None
    chart.y_axis.title = None
    chart.x_axis.title = None

    chart.legend = None  # Supprime la lÃ©gende Ã  droite

    #  Supprimer la bordure externe du graphique
    if chart.graphical_properties is None:
        chart.graphical_properties = GraphicalProperties()

    if chart.graphical_properties.line is None:
        chart.graphical_properties.line = LineProperties()

    chart.graphical_properties.line.noFill = True

    
    # Utiliser la colonne "Scholarly Output" (colonne 2) pour les donnÃ©es
    data_range = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)

    #  Define the data range for the chart labels (use Column D for Percentage Display)
    label_range = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)  # Column D (4)

    # Define the categories range for the chart (Subject Area names)
    categories_range = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)

    # Add data and categories to the chart
    chart.add_data(data_range, titles_from_data=False)  # Ensure we don't take the column title
    chart.set_categories(categories_range)

    # Configurer les Ã©tiquettes pour afficher les pourcentages
    if chart.series:
        for series in chart.series:
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True  # Afficher les valeurs
            series.dLbls.showCatName = True  # Masquer les noms des catÃ©gories
            series.dLbls.showSerName = False  # Masquer le nom de la sÃ©rie
            series.dLbls.numFmt = "0.0%"  # Formater les Ã©tiquettes en pourcentages
            series.dLbls.position = "outEnd"  # Positionner les Ã©tiquettes Ã  la fin des barres
            series.dLbls.separator = " "  #  Ensure good spacing between numbers and bars
            series.dLbls.showLeaderLines = True  #  Adds leader lines to separate labels from bars
            
            # Supprimer les petits carrÃ©s (Legend Keys)
            if series.dLbls:
                series.dLbls.showLegendKey = False

            # Configuration des barres
            series.graphicalProperties.line.noFill = True  # Supprime les bordures des barres
            series.graphicalProperties.solidFill = "1F4E79"  # Couleur de remplissage des barres
            series.spPr.ln.noFill = True  # Supprime les bordures externes

            # Supprimer les lignes de la grille principale et secondaire
            chart.x_axis.majorGridlines = None
            chart.x_axis.minorGridlines = None
            chart.y_axis.majorGridlines = None
            chart.y_axis.minorGridlines = None

    
    # Appliquer le format pourcentage avec 2 dÃ©cimales Ã  la colonne GraphPercentage
    graph_percentage_col = 4  # La 4Ã¨me colonne correspond Ã  GraphPercentage
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=graph_percentage_col, max_col=graph_percentage_col):
        for cell in row:
            cell.number_format = "0.0%"

    
    # Configurer l'axe Y pour afficher les noms des catÃ©gories
    chart.y_axis.title = None  # Masquer le titre de l'axe Y
    chart.y_axis.tickLblPos = "low"  # Positionner les Ã©tiquettes Ã  gauche des barres

    # Configuration de la mise en page
    chart.layout = Layout(manualLayout=ManualLayout(x=0.01, y=0.05, w=0.55, h=0.85))
    

    # Ajouter le graphique Ã  la feuille
    ws.add_chart(chart, "E10")

    wb.save(file_path)
    
    # Ouvrir automatiquement le fichier Excel aprÃ¨s la sauvegarde
    if os.path.exists(file_path):
        os.startfile(file_path)


    return file_path


# Interface Graphique (GUI)
class PrincDiscipApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PrincDiscip - Rapport des principales disciplines")
        self.root.geometry("800x600")
        self.root.configure(bg="#f0f0f0")

        # Logo ETS
        self.logo = tk.PhotoImage(file="ets_logo.png")
        logo_label = tk.Label(root, image=self.logo, bg="#f0f0f0")
        logo_label.pack(pady=5)

        # Titre
        title = tk.Label(root, text="Bienvenue sur PrincDiscip, le logiciel qui vous permet de gÃ©nÃ©rer automatiquement le rapport des principales disciplines des auteurs.",
                         font=("Helvetica", 12, "bold"), wraplength=750, bg="#f0f0f0")
        title.pack(pady=5)

        # Champs de saisie
        self.create_label_entry("ID Scopus de l'auteur", "author_id")
        self.create_label_entry("Nom et PrÃ©nom de l'auteur", "author_name")
        self.create_label_entry("AnnÃ©e de dÃ©but", "start_year")
        self.create_label_entry("AnnÃ©e de fin", "end_year")

        # Bouton GÃ©nÃ©rer
        self.generate_button = tk.Button(root, text="GÃ©nÃ©rer le rapport", command=self.generate_report, bg="#1f77b4", fg="white", font=("Helvetica", 10, "bold"))
        self.generate_button.pack(pady=10)
    
        # Bouton Nouvelle recherche
        self.reset_button = tk.Button(root, text="Nouvelle recherche", command=self.reset_fields, bg="#28a745", fg="white", font=("Helvetica", 10, "bold"))
        self.reset_button.pack(pady=5)


        # Zone de sortie
        self.output_text = tk.Text(root, height=10, width=90, bg="white", fg="black", wrap="word")
        self.output_text.pack(pady=5)
        self.output_text.insert("1.0", " RÃ©sultats des principales disciplines :\n")

        # Charger le chemin de sauvegarde ou demander Ã  la premiÃ¨re utilisation
        self.save_path = load_save_path()

        # Si aucun chemin n'existe, demander Ã  l'utilisateur de choisir un dossier
        if not self.save_path:
            messagebox.showinfo("PremiÃ¨re utilisation", "Veuillez choisir un dossier pour sauvegarder les rapports.")
            self.save_path = filedialog.askdirectory(title="Choisir un dossier pour les rapports")
            
            if self.save_path:
                save_save_path(self.save_path)
            else:
                messagebox.showwarning("Dossier non choisi", "Aucun dossier choisi. Les fichiers seront enregistrÃ©s sur le bureau.")
                self.save_path = os.path.join(os.path.expanduser("~"), "Desktop")


    def create_label_entry(self, label_text, attribute):
        frame = tk.Frame(self.root, bg="#f0f0f0")
        frame.pack(pady=2)
        label = tk.Label(frame, text=f"{label_text} :", font=("Helvetica", 10), bg="#f0f0f0")
        label.pack(side="left")
        entry = tk.Entry(frame, width=30)
        entry.pack(side="left", padx=10)
        setattr(self, attribute, entry)

    def generate_report(self):
        author_name = self.author_name.get().strip()
        author_id = self.author_id.get().strip()
        start_year = self.start_year.get()
        end_year = self.end_year.get()

        if not (start_year.isdigit() and end_year.isdigit()):
            messagebox.showerror("Erreur", "Veuillez entrer une plage d'annÃ©es valide.")
            return

        # Charger les clÃ©s API une seule fois
        api_key, inst_token = load_api_keys()

        # PrioritÃ© : si un nom est fourni, chercher l'ID. Sinon, utiliser l'ID.
        if author_name and not author_id:
            try:
                # Recherche de l'ID via le nom
                author_id = get_author_id_by_name(author_name, api_key, inst_token)
                self.output_text.insert("end", f" ID Scopus trouvÃ© pour {author_name}: {author_id}\n")
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de trouver un ID Scopus pour {author_name}. {str(e)}")
                return

        # Si aucun ID trouvÃ© ou saisi, arrÃªter
        if not author_id:
            messagebox.showerror("Erreur", "Veuillez entrer un ID Scopus ou un nom valide.")
            return

        #  GÃ©nÃ©ration du rapport
        try:
            total_publications = get_publication_count(author_id, int(start_year), int(end_year), api_key, inst_token)
            df_subjects = get_subject_areas(author_id, int(start_year), int(end_year), api_key, inst_token)
            file_path = create_excel_report(author_id, int(start_year), int(end_year), df_subjects, total_publications, self.save_path)

            # GÃ©nÃ©rer le rapport Excel
            file_path = create_excel_report(author_id, int(start_year), int(end_year), df_subjects, total_publications, self.save_path)

            #  Ouvrir automatiquement le fichier Excel aprÃ¨s la gÃ©nÃ©ration
            if os.path.exists(file_path):
                os.startfile(file_path)
            else:
                messagebox.showerror("Erreur", f"Le fichier {file_path} n'a pas pu Ãªtre trouvÃ©.")       
        

            #  Affichage des rÃ©sultats
            self.output_text.insert("end", "\n RÃ©sultats des principales disciplines :\n")
            for _, row in df_subjects.iterrows():
                self.output_text.insert("end", f"{row['Subject Area']}: {row['Scholarly Output']} publications ({round(row['Percentage'] * 100, 1)}%)\n")

            #  Message final
            self.output_text.insert("end", f"\n Rapport gÃ©nÃ©rÃ© avec succÃ¨s ici : {file_path}\n")

        except Exception as e:
            messagebox.showerror("Erreur", str(e))




    def reset_fields(self):
        """RÃ©initialiser les champs pour une nouvelle recherche."""
        self.author_id.delete(0, tk.END)
        self.author_name.delete(0, tk.END)  # ðŸ”¹ Vider le champ Nom et PrÃ©nom
        self.start_year.delete(0, tk.END)
        self.end_year.delete(0, tk.END)
        self.output_text.delete("1.0", tk.END)
        self.author_id.focus_set()



    

if __name__ == "__main__":
    root = tk.Tk()
    app = PrincDiscipApp(root)
    root.mainloop()
