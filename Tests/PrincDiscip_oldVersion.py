import logging
import os
import requests
import sys
import configparser
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from openpyxl.drawing.line import LineProperties


# Configuration du fichier de log
logging.basicConfig(filename="princdiscip.log", level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

# Fichier pour stocker le chemin de sauvegarde des rapports
SAVE_PATH_CONFIG = os.path.expanduser("~/.config/princdiscip_config.ini")

# Fichier pour les clés API
API_CONFIG_PATH = os.path.expanduser("~/.config/pybliometrics.cfg")



def create_complete_api_config():

    """Créer un fichier pybliometrics.cfg complet comme AutoBibPlus."""
    
    # Créer une fenêtre principale cachée pour les dialogues
    root = tk.Tk()
    root.withdraw()
    
    
    api_key = simpledialog.askstring("API Key", "Entrez votre clé API Scopus :")
    inst_token = simpledialog.askstring("Inst Token", "Entrez votre jeton d'institution (InstToken) :")

    if not api_key or not inst_token:
        messagebox.showerror("Erreur", "Les clés API sont requises pour continuer.")
        return

    config = configparser.ConfigParser()

    # Section [Authentication]
    config['Authentication'] = {
        'APIKey': api_key,
        'InstToken': inst_token
    }

    # Section [Directories] - Emplacements pour le cache
    base_cache = os.path.expanduser("~/.cache/pybliometrics")
    scopus_cache = os.path.join(base_cache, "Scopus")
    scival_cache = os.path.join(base_cache, "SciVal")

    config['Directories'] = {
        'AbstractRetrieval': os.path.join(scopus_cache, 'abstract_retrieval'),
        'AffiliationRetrieval': os.path.join(scopus_cache, 'affiliation_retrieval'),
        'AffiliationSearch': os.path.join(scopus_cache, 'affiliation_search'),
        'AuthorRetrieval': os.path.join(scopus_cache, 'author_retrieval'),
        'AuthorSearch': os.path.join(scopus_cache, 'author_search'),
        'CitationOverview': os.path.join(scopus_cache, 'citation_overview'),
        'ScopusSearch': os.path.join(scopus_cache, 'scopus_search'),
        'SerialSearch': os.path.join(scopus_cache, 'serial_search'),
        'SerialTitle': os.path.join(scopus_cache, 'serial_title'),
        'PlumXMetrics': os.path.join(scopus_cache, 'plumx'),
        'SubjectClassifications': os.path.join(scopus_cache, 'subject_classification'),
        'AuthorLookup': os.path.join(scival_cache, 'author_lookup'),
        'CountryLookup': os.path.join(scival_cache, 'country_lookup'),
        'CountryGroupLookup': os.path.join(scival_cache, 'country_group_lookup'),
        'InstitutionLookup': os.path.join(scival_cache, 'institution_lookup'),
        'InstitutionGroupLookup': os.path.join(scival_cache, 'author_group_lookup'),
        'PublicationLookup': os.path.join(scival_cache, 'publication_lookup'),
        'ScopusSourceLookup': os.path.join(scival_cache, 'scopus_source_lookup'),
        'SubjectAreaLookup': os.path.join(scival_cache, 'subject_area_lookup'),
        'TopicLookup': os.path.join(scival_cache, 'topic_lookup'),
        'TopicClusterLookup': os.path.join(scival_cache, 'topic_cluster_lookup'),
        'WorldLookup': os.path.join(scival_cache, 'world_lookup')
    }

    # Section [Requests] - Timeout et nombre de tentatives
    config['Requests'] = {
        'Timeout': '20',
        'Retries': '5'
    }

    # Section [Docs Path] - Chemin des documents générés
    user_desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    config['Docs Path'] = {
        'Path': os.path.join(user_desktop, "GraphiquesAutoBibPlus")
    }

    # Créer les répertoires si nécessaire
    os.makedirs(os.path.dirname(API_CONFIG_PATH), exist_ok=True)
    os.makedirs(scopus_cache, exist_ok=True)
    os.makedirs(scival_cache, exist_ok=True)

    # Écrire le fichier de configuration
    with open(API_CONFIG_PATH, 'w') as configfile:
        config.write(configfile)

    messagebox.showinfo("Succès", f"Le fichier de configuration complet a été créé ici : {API_CONFIG_PATH}")


def find_config_file():
    """Recherche le fichier pybliometrics.cfg dans différents emplacements."""
    possible_paths = [
        os.path.expanduser("~/.config/pybliometrics.cfg"),  # Emplacement par défaut (Linux/Mac/Windows)
        os.path.join(os.getenv("APPDATA", ""), "pybliometrics.cfg"),  # Windows : C:\Users\Nom\AppData\Roaming\
        os.path.join(os.path.expanduser("~"), ".pybliometrics.cfg"),  # Dans le dossier utilisateur
        os.path.join(os.getcwd(), "pybliometrics.cfg")  # Dossier actuel du script
    ]

    for path in possible_paths:
        if os.path.exists(path):
            return path

    return None


def load_api_keys():
    """Charger les clés API ou créer le fichier si manquant."""
    if not os.path.exists(API_CONFIG_PATH):
        logging.debug("Fichier pybliometrics.cfg introuvable. Création en cours.")
        create_complete_api_config()

    config = configparser.ConfigParser()
    config.read(API_CONFIG_PATH)

    if 'Authentication' not in config:
        logging.error("Section [Authentication] absente du fichier de configuration.")
        raise ValueError("La section [Authentication] est absente.")

    try:
        api_key = config['Authentication']['APIKey']
        inst_token = config['Authentication']['InstToken']
        logging.debug("Clés API chargées avec succès.")
        return api_key, inst_token
    except KeyError as e:
        logging.error(f"Clé manquante : {e}")
        raise ValueError(f"Clé manquante : {e}")





def load_save_path():
    """Charger le chemin de sauvegarde depuis la configuration locale."""
    config = configparser.ConfigParser()

    if os.path.exists(SAVE_PATH_CONFIG):
        config.read(SAVE_PATH_CONFIG)
        return config.get("Settings", "save_path", fallback=None)
    else:
        return None

def save_save_path(path):
    """Enregistrer le chemin de sauvegarde dans un fichier séparé."""
    config = configparser.ConfigParser()
    config["Settings"] = {"save_path": path}

    os.makedirs(os.path.dirname(SAVE_PATH_CONFIG), exist_ok=True)

    with open(SAVE_PATH_CONFIG, "w") as configfile:
        config.write(configfile)



#  Récupérer les publications par plage d'années
def get_publication_count(author_id, start_year, end_year, api_key, inst_token):
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "X-ELS-Insttoken": inst_token, "Accept": "application/json"}
    params = {"query": f"AU-ID({author_id}) AND PUBYEAR > {start_year - 1} AND PUBYEAR < {end_year + 1}", "count": "0"}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        raise RuntimeError(f"Erreur API : {response.status_code} - {response.text}")
    return int(response.json()["search-results"]["opensearch:totalResults"])

#  Récupérer les principales disciplines
def get_subject_areas(author_id, start_year, end_year, api_key, inst_token):
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "X-ELS-Insttoken": inst_token, "Accept": "application/json"}
    params = {"query": f"AU-ID({author_id}) AND PUBYEAR > {start_year-1} AND PUBYEAR < {end_year+1}", "facets": "subjarea", "count": "0"}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        raise RuntimeError(f"Erreur API : {response.status_code} - {response.text}")

    data = response.json()
    subject_areas = data.get("search-results", {}).get("facet", {}).get("category", [])
    if not subject_areas:
        raise ValueError("Aucune catégorie trouvée.")

    df = pd.DataFrame(subject_areas)
    df.rename(columns={"label": "Subject Area", "hitCount": "Scholarly Output"}, inplace=True)
    df["Scholarly Output"] = df["Scholarly Output"].astype(int)
    return df.sort_values(by="Scholarly Output", ascending=False)

def get_author_id_by_name(author_name, api_key, inst_token):
    if "," not in author_name:
        raise ValueError("Le nom doit être au format : Nom, Prénom")

    last_name, first_name = [part.strip().title() for part in author_name.split(",", 1)]
    formatted_name = f"{last_name}, {first_name}"

    url = "https://api.elsevier.com/content/search/author"
    headers = {
        "X-ELS-APIKey": api_key,
        "X-ELS-Insttoken": inst_token,
        "Accept": "application/json"
    }
    # Supprimer la virgule pour le format Nom Prénom
    formatted_name = formatted_name.replace(",", "")
    # Séparation du nom et du prénom
    last_name, first_name = formatted_name.split(" ", 1)
    params = {"query": f"AUTHLAST({last_name}) AND AUTHFIRST({first_name})", "count": 5}




    #  LOG : Afficher la requête envoyée
    print(f"\n🔍 Requête envoyée : {url}")
    print(f"Paramètres : {params}")
    print(f"En-têtes : {headers}\n")

    # Envoyer la requête
    response = requests.get(url, headers=headers, params=params)

    # LOG de la réponse brute
    print(f"Réponse brute de l'API : {response.status_code}")
    print(response.json())

    if response.status_code != 200:
        raise RuntimeError(f"Erreur API : {response.status_code} - {response.json()}")

    results = response.json().get("search-results", {}).get("entry", [])
    if not results:
        raise ValueError(f"Aucun auteur trouvé pour : {formatted_name}")

    # Afficher les auteurs trouvés
    for entry in results:
        print(f"Nom : {entry.get('preferred-name', {}).get('surname')}, {entry.get('preferred-name', {}).get('given-name')}")
        print(f"ID Scopus : {entry['dc:identifier'].split(':')[-1]}")
        print("---")

    return results[0]["dc:identifier"].split(":")[-1]





#  Créer un fichier Excel avec le graphique
def create_excel_report(author_id, start_year, end_year, df, total_publications, save_path):
    now = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    filename = f"{now}_Scopus_{author_id}.xlsx"

    # Utiliser le chemin défini par l'utilisateur pour sauvegarder le fichier
    output_dir = save_path
    os.makedirs(output_dir, exist_ok=True)
    
    # Chemin complet du fichier
    file_path = os.path.join(output_dir, filename)

    # Calcul des pourcentages corrects
    if total_publications:
        df["Percentage"] = df["Scholarly Output"] / total_publications
    else:
        df["Percentage"] = df["Scholarly Output"] / df["Scholarly Output"].sum()


    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Areas"

    #  Ajouter les informations générales en haut du tableau
    ws.append(["Data set", "Publications by Subject Area"])
    ws.append(["Entity", author_id])
    ws.append(["Subject classification", "ASJC"])
    ws.append(["Filtered by", "not filtered"])
    ws.append(["Types of publications included", "all publication types"])
    ws.append(["Self-citations", "included"])
    ws.append(["Data source", "Scopus"])
    ws.append(["Date last updated", "29 January 2025"])  #  À automatiser si nécessaire
    ws.append(["Date exported", datetime.now().strftime("%d %B %Y")])
    ws.append(["Year range", f"{start_year} to {end_year}"])  # 🔹 Afficher la plage d'années choisie
    ws.append(["Total général", total_publications])  # 🔹 Ajouter le total général
    ws.append(["Total dans la période choisie", total_publications])  # 🔹 Afficher le total dynamique
    ws.append([])  #  Ligne vide pour séparer les infos du tableau


    # Retirer "(all)" des noms des Subject Areas
    df["Subject Area"] = df["Subject Area"].str.replace(r"\s*\(all\)", "", regex=True)

   
    # Ajouter les données principales
    ws.append(["Subject Area", "Scholarly Output", "Percentage", "GraphPercentage"])  # Ajouter les en-têtes
   
   
     #  Sort DataFrame by Scholarly Output (highest to lowest)
    df = df.sort_values(by="Scholarly Output", ascending=False)
    df = df.iloc[::-1]  # Inverser l'ordre des lignes

    # Créer une colonne pour le graphique avec les pourcentages multipliés par 100 (sans affecter le tableau)
    df["GraphPercentage"] = df["Percentage"]   # Ajustement pour l'affichage correct du graphique



    for _, row in df.iterrows():
        #ws.append([row["Subject Area"], row["Scholarly Output"], row["Percentage"], round(row["GraphPercentage"], 1)])
        ws.append([row["Subject Area"], row["Scholarly Output"], row["Percentage"], row["GraphPercentage"]])
    
   
    #  Dynamically detect the starting and ending row
    start_row = ws.max_row - len(df) + 1  # Detects where data starts
    end_row = ws.max_row  # Detects last row with data

    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '0.0%'

    # Créer le graphique
    chart = BarChart()
    chart.type = "bar"  # This creates a horizontal bar chart
    chart.style = 10
    # chart.y_axis.reverseOrder = True  #  Reverse the order of bars (longest bar at the top)
    chart.y_axis.labelOffset = 0

    chart.title = None
    chart.y_axis.title = None
    chart.x_axis.title = None

    chart.legend = None  # Supprime la légende à droite

    #  Supprimer la bordure externe du graphique
    if chart.graphical_properties is None:
        chart.graphical_properties = GraphicalProperties()

    if chart.graphical_properties.line is None:
        chart.graphical_properties.line = LineProperties()

    chart.graphical_properties.line.noFill = True

    
    # Utiliser la colonne "Scholarly Output" (colonne 2) pour les données
    data_range = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)

    #  Define the data range for the chart labels (use Column D for Percentage Display)
    label_range = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)  # Column D (4)

    # Define the categories range for the chart (Subject Area names)
    categories_range = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)

    # Add data and categories to the chart
    chart.add_data(data_range, titles_from_data=False)  # Ensure we don't take the column title
    chart.set_categories(categories_range)

    # Configurer les étiquettes pour afficher les pourcentages
    if chart.series:
        for series in chart.series:
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True  # Afficher les valeurs
            series.dLbls.showCatName = True  # Masquer les noms des catégories
            series.dLbls.showSerName = False  # Masquer le nom de la série
            series.dLbls.numFmt = "0.0%"  # Formater les étiquettes en pourcentages
            series.dLbls.position = "outEnd"  # Positionner les étiquettes à la fin des barres
            series.dLbls.separator = " "  #  Ensure good spacing between numbers and bars
            series.dLbls.showLeaderLines = True  #  Adds leader lines to separate labels from bars
            
            # Supprimer les petits carrés (Legend Keys)
            if series.dLbls:
                series.dLbls.showLegendKey = False

            # Configuration des barres
            series.graphicalProperties.line.noFill = True  # Supprime les bordures des barres
            series.graphicalProperties.solidFill = "1F4E79"  # Couleur de remplissage des barres
            series.spPr.ln.noFill = True  # Supprime les bordures externes

            # Supprimer les lignes de la grille principale et secondaire
            chart.x_axis.majorGridlines = None
            chart.x_axis.minorGridlines = None
            chart.y_axis.majorGridlines = None
            chart.y_axis.minorGridlines = None

    
    # Appliquer le format pourcentage avec 2 décimales à la colonne GraphPercentage
    graph_percentage_col = 4  # La 4ème colonne correspond à GraphPercentage
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=graph_percentage_col, max_col=graph_percentage_col):
        for cell in row:
            cell.number_format = "0.0%"

    
    # Configurer l'axe Y pour afficher les noms des catégories
    chart.y_axis.title = None  # Masquer le titre de l'axe Y
    chart.y_axis.tickLblPos = "low"  # Positionner les étiquettes à gauche des barres

    # Configuration de la mise en page
    chart.layout = Layout(manualLayout=ManualLayout(x=0.01, y=0.05, w=0.55, h=0.85))
    

    # Ajouter le graphique à la feuille
    ws.add_chart(chart, "E10")

    wb.save(file_path)
    
    # # Ouvrir automatiquement le fichier Excel après la sauvegarde
    # if os.path.exists(file_path):
    #     os.startfile(file_path)


    return file_path


# Interface Graphique (GUI)
class PrincDiscipApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PrincDiscip - Rapport des principales disciplines")
        self.root.geometry("800x600")
        self.root.configure(bg="#f0f0f0")

        
        # Charger le logo ETS en gérant le chemin pour PyInstaller
        if getattr(sys, 'frozen', False):
            # Si exécuté en tant qu'exécutable
            base_path = sys._MEIPASS
        else:
            # Si exécuté comme script Python
            base_path = os.path.abspath(".")

        # Chemin de l'image du logo
        logo_path = os.path.join(base_path, "ETS_logo.png")

        # Vérification de l'existence de l'image
        if not os.path.exists(logo_path):
            messagebox.showerror("Erreur", f"L'image du logo est introuvable : {logo_path}")

        # Charger l'image du logo
        self.logo = tk.PhotoImage(file=logo_path)
        logo_label = tk.Label(root, image=self.logo, bg="#f0f0f0")
        logo_label.pack(pady=5)


        # Titre
        title = tk.Label(root, text="Bienvenue sur PrincDiscip, le logiciel qui vous permet de générer automatiquement le rapport des principales disciplines des auteurs.",
                         font=("Helvetica", 12, "bold"), wraplength=750, bg="#f0f0f0")
        title.pack(pady=5)

        # Champs de saisie
        self.create_label_entry("ID Scopus de l'auteur", "author_id")
        self.create_label_entry("Nom et Prénom de l'auteur", "author_name")
        self.create_label_entry("Année de début", "start_year")
        self.create_label_entry("Année de fin", "end_year")

        # Variable pour la case à cocher
        self.use_default_years = tk.BooleanVar()

        # Case à cocher pour activer la plage par défaut des 5 dernières années
        self.default_years_check = tk.Checkbutton(
            root,
            text="Utiliser les 5 dernières années",
            variable=self.use_default_years,
            command=self.toggle_years,
            bg="#f0f0f0"
        )
        self.default_years_check.pack(pady=2)


        # Bouton Générer
        self.generate_button = tk.Button(root, text="Générer le rapport", command=self.generate_report, bg="#1f77b4", fg="white", font=("Helvetica", 10, "bold"))
        self.generate_button.pack(pady=10)
    
        # Bouton Nouvelle recherche
        self.reset_button = tk.Button(root, text="Nouvelle recherche", command=self.reset_fields, bg="#28a745", fg="white", font=("Helvetica", 10, "bold"))
        self.reset_button.pack(pady=5)


        # Zone de sortie
        self.output_text = tk.Text(root, height=10, width=90, bg="white", fg="black", wrap="word")
        self.output_text.pack(pady=5)
        self.output_text.insert("1.0", " Résultats des principales disciplines :\n")

        # Charger le chemin de sauvegarde ou demander à la première utilisation
        self.save_path = load_save_path()

        # Si aucun chemin n'existe, demander à l'utilisateur de choisir un dossier
        if not self.save_path:
            messagebox.showinfo("Première utilisation", "Veuillez choisir un dossier pour sauvegarder les rapports.")
            self.save_path = filedialog.askdirectory(title="Choisir un dossier pour les rapports")
            
            if self.save_path:
                save_save_path(self.save_path)
            else:
                messagebox.showwarning("Dossier non choisi", "Aucun dossier choisi. Les fichiers seront enregistrés sur le bureau.")
                self.save_path = os.path.join(os.path.expanduser("~"), "Desktop")


    def create_label_entry(self, label_text, attribute):
        frame = tk.Frame(self.root, bg="#f0f0f0")
        frame.pack(pady=2)
        label = tk.Label(frame, text=f"{label_text} :", font=("Helvetica", 10), bg="#f0f0f0")
        label.pack(side="left")
        entry = tk.Entry(frame, width=30)
        entry.pack(side="left", padx=10)
        setattr(self, attribute, entry)

    
    def toggle_years(self):
        """Activer ou désactiver les champs d'année selon la case cochée."""
        if self.use_default_years.get():
            current_year = datetime.now().year
            self.start_year.delete(0, tk.END)
            self.end_year.delete(0, tk.END)
            
            # Plage correcte : 5 dernières années, y compris l'année en cours
            self.start_year.insert(0, str(current_year - 6))  # 2019 si on est en 2025
            self.end_year.insert(0, str(current_year))        # Inclure 2025
            
            # Désactiver les champs pour éviter la saisie manuelle
            self.start_year.config(state='disabled')
            self.end_year.config(state='disabled')
        else:
            # Réactiver la saisie manuelle si décoché
            self.start_year.config(state='normal')
            self.end_year.config(state='normal')



    def generate_report(self):
        author_name = self.author_name.get().strip()
        author_id = self.author_id.get().strip()

        logging.debug("Début de la génération du rapport.")
        try:
            # Déterminer la plage d'années
            if self.use_default_years.get():
                current_year = datetime.now().year
                start_year = current_year - 6
                end_year = current_year
            else:
                start_year = int(self.start_year.get())
                end_year = int(self.end_year.get())

            logging.debug(f"Plage d'années : {start_year} - {end_year}")

            # Charger les clés API
            api_key, inst_token = load_api_keys()

            # Rechercher l'ID Scopus si le nom est fourni
            if author_name and not author_id:
                author_id = get_author_id_by_name(author_name, api_key, inst_token)
                logging.debug(f"ID Scopus trouvé : {author_id}")

            if not author_id:
                logging.error("Aucun ID Scopus fourni.")
                messagebox.showerror("Erreur", "Veuillez entrer un ID Scopus valide.")
                return

            # Récupérer les données
            logging.debug(f"Récupération des publications pour l'ID : {author_id}")
            total_publications = get_publication_count(author_id, start_year, end_year, api_key, inst_token)
            logging.debug(f"Total des publications : {total_publications}")

            # Générer le rapport Excel
            logging.debug(f"Chemin de sauvegarde : {self.save_path}")
            df_subjects = get_subject_areas(author_id, start_year, end_year, api_key, inst_token)
            file_path = create_excel_report(author_id, start_year, end_year, df_subjects, total_publications, self.save_path)

            # Afficher le message de succès et attendre la fermeture
            if messagebox.showinfo("Succès", f"Rapport généré avec succès !\n{file_path}") == "ok":
                # Ouvrir le fichier Excel seulement après la fermeture du message
                if os.path.exists(file_path):
                    os.startfile(file_path)

            else:
                logging.error("Le fichier Excel n'a pas été créé.")
                messagebox.showerror("Erreur", "Le fichier Excel n'a pas été créé.")
        
        except Exception as e:
            logging.error(f"Erreur durant la génération : {e}")
            messagebox.showerror("Erreur critique", f"Une erreur est survenue : {str(e)}")




    def reset_fields(self):
        """Réinitialiser les champs pour une nouvelle recherche."""
        # Vider tous les champs
        self.author_id.delete(0, tk.END)
        self.author_name.delete(0, tk.END)
        self.start_year.delete(0, tk.END)
        self.end_year.delete(0, tk.END)
        
        # Décocher la case par défaut
        self.use_default_years.set(False)

        # Activer les champs d'année et les vider
        self.start_year.config(state='normal')
        self.end_year.config(state='normal')
        self.start_year.delete(0, tk.END)
        self.end_year.delete(0, tk.END)

        # Effacer la zone de sortie
        self.output_text.delete("1.0", tk.END)
        self.author_id.focus_set()



    

if __name__ == "__main__":
    root = tk.Tk()
    app = PrincDiscipApp(root)
    root.mainloop()
