import os
import openpyxl
import requests
import configparser
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime

from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties

from collections import defaultdict
import time



# ✅ Fonction pour récupérer les sous-catégories de chaque publication de l’auteur
def get_subcategories(AUTHOR_ID, start_year, end_year, headers):
    subcategory_data = defaultdict(int)
    subcategory_asjc_codes = {}
    search_url = "https://api.elsevier.com/content/search/scopus"
    abstract_url = "https://api.elsevier.com/content/abstract/scopus_id/"
    
    # ✅ Étape 1: Obtenir les IDs des documents
    batch_size = 10
    offset = 0
    all_scopus_ids = []

    while True:
        params = {
            "query": f"AU-ID({AUTHOR_ID}) AND PUBYEAR > {start_year-1} AND PUBYEAR < {end_year+1}",
            "count": str(batch_size),
            "start": str(offset),
            "view": "COMPLETE"
        }

        response = requests.get(search_url, headers=headers, params=params)

        if response.status_code != 200:
            print(f"❌ Erreur API Scopus : {response.status_code} - {response.text}")
            break

        data = response.json()
        entries = data.get("search-results", {}).get("entry", [])

        if not entries:
            print("✅ Plus aucun document trouvé. Arrêt du traitement.")
            break

        batch_scopus_ids = [entry.get("dc:identifier", "").replace("SCOPUS_ID:", "").strip() for entry in entries if entry.get("dc:identifier")]

        if not batch_scopus_ids:
            print("⚠️ Aucun SCOPUS_ID trouvé dans ce lot.")
            break

        all_scopus_ids.extend(batch_scopus_ids)
        offset += batch_size
        time.sleep(2)  # ✅ Pause pour éviter les limites de requêtes

    # ✅ Étape 2: Récupérer les sous-catégories pour chaque publication
    seen_subcategories = set()  # ✅ Pour éviter les doublons

    for index, scopus_id in enumerate(all_scopus_ids, start=1):
        print(f"🔍 Récupération des sous-catégories pour SCOPUS_ID {scopus_id} ({index}/{len(all_scopus_ids)})...")

        abstract_response = requests.get(f"{abstract_url}{scopus_id}", headers=headers)

        if abstract_response.status_code == 429:  # Erreur "Trop de requêtes"
            print("⚠️ Limite atteinte. Pause de 10 secondes...")
            time.sleep(10)
            continue

        if abstract_response.status_code != 200:
            print(f"⚠️ Échec de récupération pour SCOPUS_ID {scopus_id}. Passage au suivant...")
            continue

        abstract_data = abstract_response.json()
        subject_areas = abstract_data.get("abstracts-retrieval-response", {}).get("subject-areas", {}).get("subject-area", [])

        if isinstance(subject_areas, list):
            for subject in subject_areas:
                subject_name = subject.get("$", "").strip()
                asjc_code = subject.get("@code", "N/A")

                # ✅ Éviter les doublons
                if (subject_name, asjc_code) not in seen_subcategories:
                    seen_subcategories.add((subject_name, asjc_code))
                    subcategory_data[subject_name] += 1
                    subcategory_asjc_codes[subject_name] = asjc_code

        time.sleep(2)  # ✅ Pause pour éviter les limitations d'API


    # ✅ Étape 3: Structurer les résultats
    subcategory_results = [
        {
            "ASJC Code": subcategory_asjc_codes.get(subject, "N/A"),
            "Category": subject,
            "Scholarly Output": count
        }
        for subject, count in subcategory_data.items()
    ]

    return subcategory_results




now = datetime.now().strftime("%Y-%m-%d_%Hh%M")  # ✅ Définition 


# 📌 Charger la configuration depuis pybliometrics.cfg
config = configparser.ConfigParser()
config.read(r"C:\Users\alima\.config\pybliometrics.cfg") # Sur Windows
# config.read(r"/Users/alexandreleao/.config/pybliometrics.cfg") # Sur MacOs

API_KEY = config["Authentication"]["APIKey"]
INST_TOKEN = config["Authentication"]["InstToken"]

# ✅ Vérifier si l'API Key et le Token sont bien récupérés
if not API_KEY or not INST_TOKEN:
    print("\u274c Erreur : Impossible de récupérer l'API Key ou le Token Institutionnel depuis pybliometrics.cfg")
    exit()

print(f"\u2705 API Key récupérée : {API_KEY[:5]}**")
print(f"\u2705 Institution Token récupéré : {INST_TOKEN[:5]}**\n")

# ✅ Demander l'ID de l’auteur
AUTHOR_ID = input("Entrez l'ID Scopus de l’auteur : ").strip()

# ✅ Obtenir l'année actuelle
current_year = datetime.now().year

# ✅ Demander à l'utilisateur de saisir la plage d'années avec une option par défaut
start_year_input = input(f"Entrez l'année de début (ou appuyez sur Entrée pour {current_year - 6}) : ").strip()
end_year_input = input(f"Entrez l'année de fin (ou appuyez sur Entrée pour {current_year}) : ").strip()

# ✅ Vérifier et attribuer les valeurs par défaut si l'utilisateur appuie sur Entrée
if start_year_input and not start_year_input.isdigit():
    print("\u274c Erreur : L'année de début doit être un nombre entier.")
    exit()
if end_year_input and not end_year_input.isdigit():
    print("\u274c Erreur : L'année de fin doit être un nombre entier.")
    exit()

start_year = int(start_year_input) if start_year_input else current_year - 6
end_year = int(end_year_input) if end_year_input else current_year

# ✅ Vérifier que l'année de fin est postérieure ou égale à l'année de début
if start_year > end_year:
    print("\u274c Erreur : L'année de début doit être inférieure ou égale à l'année de fin.")
    exit()

print(f"📊 Plage d'années sélectionnée : {start_year} - {end_year}")




if not AUTHOR_ID:
    print("\u274c Erreur : L'ID de l'auteur ne peut pas être vide.")
    exit()

# ✅ Définition des headers AVANT d'effectuer le premier appel API
headers = {
    "X-ELS-APIKey": API_KEY,
    "X-ELS-Insttoken": INST_TOKEN,
    "Accept": "application/json"
}

# ✅ Définition de l'URL de l'API Scopus pour récupérer le total des publications sur la plage choisie
SCOPUS_TOTAL_URL = "https://api.elsevier.com/content/search/scopus"
params_total = {
    "query": f"AU-ID({AUTHOR_ID}) AND PUBYEAR > {start_year-1} AND PUBYEAR < {end_year+1}",
    "count": "0"
}



response_total = requests.get(SCOPUS_TOTAL_URL, headers=headers, params=params_total)

if response_total.status_code == 200:
    total_data = response_total.json()
    try:
        total_publications_plage = int(total_data["search-results"]["opensearch:totalResults"])
        print(f"\u2705 Nombre total de publications ({start_year}-{end_year}) : {total_publications_plage}")
    except KeyError:
        print("\u26a0\ufe0f Impossible de récupérer le nombre total de publications sur la période demandée.")
        total_publications_plage = None
else:
    print(f"\u274c Erreur API Scopus : {response_total.status_code} - {response_total.text}")
    total_publications_plage = None




response_total = requests.get(SCOPUS_TOTAL_URL, headers=headers, params=params_total)

if response_total.status_code == 200:
    total_data = response_total.json()
    try:
        total_publications = int(total_data["search-results"]["opensearch:totalResults"])
        print(f"\u2705 Nombre total de publications (2019-2025) : {total_publications}")
    except KeyError:
        print("\u26a0\ufe0f Impossible de récupérer le nombre total de publications. Utilisation de la somme des Subject Areas.")
        total_publications = None
else:
    print(f"\u274c Erreur API Scopus : {response_total.status_code} - {response_total.text}")
    total_publications = None



# ✅ Définition de l'URL de l'API Scopus pour les Subject Areas
params = {
    "query": f"AU-ID({AUTHOR_ID}) AND PUBYEAR > 2018 AND PUBYEAR < 2026",
    "facets": "subjarea",
    "count": "0"
}

response = requests.get(SCOPUS_TOTAL_URL, headers=headers, params=params)

if response.status_code != 200:
    print(f"\u274c Erreur API Scopus : {response.status_code} - {response.text}")
    exit()

data = response.json()

if "search-results" in data and "facet" in data["search-results"]:
    subject_areas = data["search-results"]["facet"]["category"]
    df = pd.DataFrame(subject_areas)
    df = df.rename(columns={"label": "Subject Area", "hitCount": "Scholarly Output"})
    df["Scholarly Output"] = df["Scholarly Output"].astype(int)

    
    # Trier par ordre décroissant
    df = df.sort_values(by="Scholarly Output", ascending=False)

    # ✅ Removing "(all)" from subject names
    df["Subject Area"] = df["Subject Area"].str.replace(r"\s*\(all\)", "", regex=True)

    choix_utilisateur = ""  # ✅ Initialisation pour éviter l'erreur NameError

    # ✅ Afficher les principales catégories et sous-catégories à l'utilisateur
    print("\n📌 Voici les principales catégories disponibles :")
    print(df[["Subject Area", "Scholarly Output"]].to_string(index=False))

    # ✅ Récupérer les sous-catégories avant de les afficher
    subcategories = get_subcategories(AUTHOR_ID, start_year, end_year, headers)


    print("\n📌 Voici les sous-catégories disponibles :")
    print(pd.DataFrame(subcategories, columns=["ASJC Code", "Category", "Scholarly Output"]).to_string(index=False))

    # ✅ Demander à l'utilisateur s'il souhaite choisir manuellement ou utiliser la sélection par défaut
    user_input = input("\nSouhaitez-vous choisir vous-même les catégories/sous-catégories ? (oui/non) : ").strip().lower()

    if user_input == "oui":
        choix_utilisateur = input("\nEntrez les 5 codes ASJC ou noms des catégories/sous-catégories séparés par une virgule : ").strip()
    else:
        print("\n✅ Utilisation des 5 principales catégories par défaut.")
        choix_utilisateur = "default"  # 🔹 Ajoutez cette ligne pour éviter l'erreur


    # ✅ Demander à l’utilisateur s'il souhaite choisir manuellement ou utiliser les 5 principales par défaut
    mode_choix = input("\nSouhaitez-vous choisir vous-même les catégories/sous-catégories ? (oui/non) : ").strip().lower()

    if mode_choix == "oui":
        choix_utilisateur = input("\nEntrez les 5 codes ASJC ou noms des catégories/sous-catégories séparés par une virgule : ").strip()
        choix_codes = [code.strip() for code in choix_utilisateur.split(",")]
        choix_final = df[df["Subject Area"].isin(choix_codes) | df["ASJC Code"].astype(str).isin(choix_codes)]
        
        if choix_final.empty:
            print("⚠️ Aucun élément correspondant trouvé. On garde la sélection par défaut.")
            choix_final = df.head(5)
    else:
        print("\n✅ Utilisation des 5 principales catégories par défaut.")
        choix_final = df.head(5)

    choix_utilisateur = "default"

    print(f"DEBUG: choix_utilisateur = '{choix_utilisateur}'")  # 🔍 Vérifier si la variable existe bien


    # ✅ Si l'utilisateur ne choisit rien, prendre les 5 premières catégories principales
    if not choix_utilisateur:
        choix_final = df.head(5)  # Prend les 5 principales catégories par défaut
    else:
        choix_codes = [code.strip() for code in choix_utilisateur.split(",")]

        print(f"DEBUG: Colonnes disponibles dans df: {df.columns.tolist()}")

        if "ASJC Code" in df.columns:
            choix_final = df[df["Subject Area"].isin(choix_codes) | df["ASJC Code"].astype(str).isin(choix_codes)]
        else:
            choix_final = df[df["Subject Area"].isin(choix_codes)]  # ✅ Si "ASJC Code" n'existe pas, on l'ignore

        
        if choix_final.empty:
            print("⚠️ Aucun élément correspondant trouvé. On garde la sélection par défaut.")
            choix_final = df.head(5)  # Sélectionne les 5 premières catégories si l'utilisateur ne choisit rien
 
    # ✅ Afficher les catégories et sous-catégories sélectionnées
    print("\n✅ Catégories et sous-catégories sélectionnées pour le rapport :")
    print(choix_final)


    # Calcul des pourcentages corrects
    if total_publications:
        df["Percentage"] = (df["Scholarly Output"] / total_publications) * 100
    else:
        df["Percentage"] = (df["Scholarly Output"] / df["Scholarly Output"].sum()) * 100


    # ✅ Récupérer les sous-catégories
    subcategories = get_subcategories(AUTHOR_ID, start_year, end_year, headers)

    # ✅ Afficher les 5 principales catégories avec leurs Scholarly Outputs
    print("\n🔹 Les 5 principales catégories par défaut :")
    top_5_categories = df.nlargest(5, "Scholarly Output")[["Subject Area", "Scholarly Output"]]
    print(top_5_categories.to_string(index=False))

    # ✅ Afficher les sous-catégories récupérées
    print("\n🔹 Sous-catégories disponibles :")
    df_subcategories = pd.DataFrame(subcategories)
    df_subcategories = df_subcategories.sort_values(by="Scholarly Output", ascending=False)
    print(df_subcategories.to_string(index=False))

    # ✅ Demander à l'utilisateur de choisir ses catégories/sous-catégories
    user_input = input("\nSouhaitez-vous modifier cette sélection ? (oui/non) : ").strip().lower()

    if user_input == "oui":
        print("\nEntrez les noms des catégories/sous-catégories que vous souhaitez inclure (séparés par des virgules).")
        print("Exemple : Engineering, Business, Architecture, Civil and Structural Engineering, Tourism, etc.")

        user_selected_names = input("\n👉 Entrez vos choix ici : ").strip()

        if user_selected_names:
            selected_names = [name.strip() for name in user_selected_names.split(",") if name.strip()]
            
            # ✅ Filtrer les catégories et sous-catégories sélectionnées par l'utilisateur
            df_selected_categories = df[df["Subject Area"].isin(selected_names)]
            df_selected_subcategories = df_subcategories[df_subcategories["Category"].isin(selected_names)]
            
            if df_selected_categories.empty and df_selected_subcategories.empty:
                print("⚠️ Aucun élément correspondant trouvé. On garde la sélection par défaut.")
            else:
                top_5_categories = df_selected_categories if not df_selected_categories.empty else top_5_categories
                subcategories = df_selected_subcategories.to_dict(orient="records")

    # ✅ Afficher la sélection finale
    print("\n✅ Catégories/Sous-catégories sélectionnées pour le fichier Excel et le graphique :")
    print(top_5_categories.to_string(index=False))
    print(df_subcategories.to_string(index=False))


    # 📌 Récupérer le nom et prénom de l'auteur AVANT de générer le nom du fichier
    SCOPUS_AUTHOR_URL = f"https://api.elsevier.com/content/author/author_id/{AUTHOR_ID}"

    response = requests.get(SCOPUS_AUTHOR_URL, headers=headers)

    if response.status_code == 200:
        author_data = response.json()  # 📌 Définition correcte de author_data
        try:
            total_publications = int(author_data["author-retrieval-response"][0]["coredata"]["document-count"])
            print(f"✅ Nombre total de publications (2019-2025) : {total_publications}")
        except KeyError:
            print("⚠️ Impossible de récupérer le nombre total de publications. Utilisation de la somme des Subject Areas.")
            total_publications = None
    else:
        print(f"❌ Erreur API Scopus : {response.status_code} - {response.text}")
        total_publications = None



    # ✅ Récupérer le nom et prénom de l'auteur AVANT de générer le nom du fichier
    try:
        author_name = author_data["author-retrieval-response"][0]["author-profile"]["preferred-name"]
        author_firstname = author_name["given-name"]
        author_lastname = author_name["surname"]
    except KeyError:
        print("⚠️ Impossible de récupérer le nom complet de l’auteur. Utilisation de l’ID.")
        author_firstname = "Auteur"
        author_lastname = AUTHOR_ID

    # ✅ Générer le nom du fichier après avoir défini les variables
    filename = f"{now}{author_lastname}{author_firstname.replace(' ', '_')}.xlsx"


    # ✅ Exporter vers Excel
    now = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    filename = f"{now}{author_lastname}{author_firstname.replace(' ', '_')}.xlsx"

    file_path = os.path.join(r"C:\Users\alima\Desktop\AlexStage\Code_graphique_test_project\output", filename) # Sur Windows
    # file_path = os.path.join(r"/Users/alexandreleao/Documents/projectBiblio/output", filename) # Sur MacOs
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Areas"

    # ✅ Ajouter les informations générales en haut du tableau
    ws.append(["Data set", "Publications by Subject Area"])
    ws.append(["Entity", AUTHOR_ID])
    ws.append(["Subject classification", "ASJC"])
    ws.append(["Filtered by", "not filtered"])
    ws.append(["Types of publications included", "all publication types"])
    ws.append(["Self-citations", "included"])
    ws.append(["Data source", "Scopus"])
    ws.append(["Date last updated", "29 January 2025"])  # 📌 À automatiser si nécessaire
    ws.append(["Date exported", datetime.now().strftime("%d %B %Y")])
    ws.append(["Year range", f"{start_year} to {end_year}"])  # 🔹 Afficher la plage d'années choisie
    ws.append(["Total général", total_publications])  # 🔹 Ajouter le total général
    ws.append(["Total dans la période choisie", total_publications_plage])  # 🔹 Afficher le total dynamique
    ws.append([])  # ✅ Ligne vide pour séparer les infos du tableau


    # ✅ Ensure only the required columns are kept
    df = df[["Subject Area", "Scholarly Output", "Percentage"]]

    # ✅ Convert Percentage column to numeric and clean up formatting
    df["Percentage"] = pd.to_numeric(df["Percentage"], errors="coerce")

    # ✅ Add a new column for displaying percentages with "%" symbol
    df["Percentage Display"] = df["Percentage"].apply(lambda x: f"{x:.1f}%")

    # ✅ Format numbers to 1 decimal place
    df["Scholarly Output"] = df["Scholarly Output"].round(1)
    df["Percentage"] = df["Percentage"].round(1)

    # ✅ Sort DataFrame by Scholarly Output (highest to lowest)
    df = df.sort_values(by="Scholarly Output", ascending=False)

    

    # ✅ Import nécessaire AVANT d'utiliser DataLabelList
    from openpyxl.chart.label import DataLabelList  



# ✅ Dynamically detect the starting and ending row
start_row = ws.max_row - len(df) + 1  # Detects where data starts
end_row = ws.max_row  # Detects last row with data


# Define the data range for the chart (Percentage column)
data_range = Reference(ws, min_col=3, min_row=start_row, max_row=end_row)  # Column C (3)

# ✅ Define the data range for the chart labels (use Column D for Percentage Display)
label_range = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)  # Column D (4)

# Define the categories range for the chart (Subject Area names)
categories_range = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)

from openpyxl.chart.shapes import GraphicalProperties

# Create a horizontal bar chart
chart = BarChart()
chart.type = "bar"  # Horizontal bar chart
chart.title = None
chart.y_axis.title = None
chart.x_axis.title = None

# ✅ Properly remove the external grey frame (border)
if chart.graphical_properties is None:
    chart.graphical_properties = GraphicalProperties()

chart.graphical_properties.line.noFill = True  # ✅ Ensures no border is applied

# Add data and categories to the chart
chart.add_data(data_range, titles_from_data=False)  # Ensure we don't take the column title
chart.set_categories(categories_range)

if chart.series:
    for series in chart.series:
        if series is not None:  # ✅ Ensure series exists before modifying
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True  # ✅ Show percentage values correctly (now from Column D)
            series.dLbls.showCatName = True  # ❌ Hide category names from labels
            series.dLbls.showSerName = False  # ✅ Disable "Series1"
            series.dLbls.position = "outEnd"  # ✅ Position labels at the end of bars
            series.dLbls.numFmt = "0.0%"  # ✅ Ensure 1 decimal place + "%"
            series.dLbls.separator = " "  # ✅ Ensure good spacing between numbers and bars
            series.dLbls.showLeaderLines = True  # ✅ Adds leader lines to separate labels from bars

            # ✅ Correctly set label color to deep blue
            series.dLbls.txPr = RichText(
                p=[ParagraphProperties(defRPr=CharacterProperties(solidFill="1F4E79"))]  # ✅ Dark blue labels
            )



    # ✅ Reduce bar width to make it thinner
    series.gapWidth = 400  # Increase gap width (default is ~150)

    # ✅ Simulate rounded markers by adding an extra small bar
    # ✅ Remove ALL border lines (internal & external)
    series.graphicalProperties.line.noFill = True  # Removes outline
    series.graphicalProperties.solidFill = "1F4E79"  # Keep solid blue fill
    series.spPr.ln.noFill = True  # ✅ Ensures NO external border is applied

    
    
    # ✅ Remove the "Series1" legend
    chart.legend = None  # Completely remove the legend

    # # ✅ Restore bar colors
    # for series in chart.series:
    #     series.graphicalProperties.solidFill = "4472C4"  # Restore blue color

    # ✅ Reverse the Y-axis order to invert labels
    chart.y_axis.reverseOrder = True  # Invert label order on the graph
    chart.y_axis.tickLblPos = "low"  # Align labels properly

    # Style the chart
    chart.style = 10  # Apply a predefined style
    chart.y_axis.majorGridlines = None  # Remove gridlines
    chart.x_axis.majorGridlines = None  # Remove gridlines

    chart.y_axis.majorTickMark = "none"  # ✅ Removes extra tick marks
    chart.y_axis.labelOffset = 0  # ✅ Moves the category labels closer to the bars


    chart.layout = Layout(
    manualLayout=ManualLayout(
        x=0.01,  # Move bars even closer to left
        y=0.05,
        w=0.7,  # Adjust width for better spacing
        h=0.85
    )
)



# ✅ Créer une feuille Excel pour enregistrer les choix de l'utilisateur
ws_selected = wb.create_sheet(title="Choix Utilisateur")

# ✅ Mettre à jour le titre de la colonne
ws_selected.append(["Subject Area/Subcategory", "Scholarly Output", "Percentage Display"])

# ✅ Ajouter les choix de l'utilisateur
for _, row in choix_final.iterrows():
    ws_selected.append(row.tolist())

# ✅ Ajouter les sous-catégories choisies par l'utilisateur
ws_subcategories = wb.create_sheet(title="Subcategories Choisies")
ws_subcategories.append(["ASJC Code", "Subcategory", "Scholarly Output"])
for sub in subcategories:
  if sub["Category"] in choix_final["Subject Area"].values or ("ASJC Code" in choix_final.columns and sub["ASJC Code"] in choix_final["ASJC Code"].astype(str).values):

        ws_subcategories.append([sub["ASJC Code"], sub["Category"], sub["Scholarly Output"]])



# ✅ Vérifier qu'il y a suffisamment de données pour un graphique
if len(choix_final) > 1:
    # ✅ Détection des lignes pour le graphique
    start_row = ws_selected.max_row - len(choix_final) + 1
    end_row = ws_selected.max_row

    # ✅ Définir les plages de données pour le graphique
    data_range = Reference(ws_selected, min_col=2, min_row=start_row, max_row=end_row)  # Scholarly Output
    categories_range = Reference(ws_selected, min_col=1, min_row=start_row, max_row=end_row)  # Nom des catégories

    # ✅ Création du graphique
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top 5 Selected Categories/Subcategories"
    chart.add_data(data_range, titles_from_data=False)
    chart.set_categories(categories_range)
    chart.y_axis.reverseOrder = True  # Inverser l'ordre des labels

    # ✅ Ajouter le graphique à l'Excel
    ws_selected.add_chart(chart, "E10")
else:
    print("⚠️ Pas assez de données pour un graphique.")


# ✅ Sauvegarde du fichier Excel
wb.save(file_path)
print(f"\n✅ Fichier Excel généré avec les choix de l'utilisateur : {filename}")
print("\n🎉 Le fichier Excel a été sauvegardé avec succès. Vous pouvez l'ouvrir maintenant !")
